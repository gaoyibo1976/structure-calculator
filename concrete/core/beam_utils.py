# -*- coding: utf-8 -*-
"""
混凝土梁抗弯承载力计算工具函数模块
"""
import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from ..config import OUTPUT_COLS, COL_MAPPING


def validate_file_exists(file_path):
    """
    验证文件是否存在
    :param file_path: 文件路径
    :raises FileNotFoundError: 当文件不存在时抛出异常
    """
    if not os.path.exists(file_path):
        print(f"❌ 未找到Excel文件！请确认文件路径：{file_path}")
        sys.exit()


def read_excel_data(file_path):
    """
    读取Excel A-P列数据
    :param file_path: Excel文件路径
    :return: pandas.DataFrame - 读取的数据
    :raises Exception: 当读取文件失败时抛出异常
    """
    try:
        df_input = pd.read_excel(
            file_path,
            sheet_name="Sheet1",
            usecols=["截面编号", "截面类型", "b", "h", "bf", "hf",
                     "混凝土强度等级C", "受拉钢筋强度等级", "受压钢筋强度等级",
                     "受拉钢筋面积As", "受拉钢筋as", "受压钢筋面积As", "受压钢筋as",
                     "弯矩设计值M", "是否地震作用组合", "结构重要性系数γ0"],
            engine="openpyxl",
            dtype={"截面编号": str}
        )
        return df_input
    except Exception as e:
        print(f"❌ 读取Excel文件时出错: {e}")
        sys.exit()


def prepare_calculation_data(df_input):
    """
    准备计算数据
    :param df_input: 输入数据DataFrame
    :return: tuple - (计算参数列表, 结果数据列表)
    """
    param = []
    result_data = []

    for _, row in df_input.iterrows():
        # 初始化结果数据
        result_item = row.to_dict()
        for col in OUTPUT_COLS.values():
            result_item[col] = None
        result_data.append(result_item)

        # 构造计算参数
        param_item = {
            "sec_num": row["截面编号"],
            "sec_type": row["截面类型"],
            "M": row["弯矩设计值M"],
            "is_seismic": row["是否地震作用组合"],
            "γ0": row["结构重要性系数γ0"],
            "calc_params": [
                row["b"], row["h"], row["bf"], row["hf"],
                row["混凝土强度等级C"], row["受拉钢筋强度等级"], row["受压钢筋强度等级"],
                row["受拉钢筋面积As"], row["受拉钢筋as"],
                row["受压钢筋面积As"], row["受压钢筋as"], row["结构重要性系数γ0"]
            ]
        }
        param.append(param_item)

    return param, result_data


def save_excel_result_with_style(result_list, save_path, source_path):
    """
    保存Excel结果，统一设置样式：数字类型、居中对齐
    :param result_list: 结果数据列表
    :param save_path: 保存路径
    :param source_path: 源文件路径
    """
    # 1. 加载原始Excel文件，保留所有样式
    wb = load_workbook(source_path)
    ws = wb.active

    # 2. 定义统一的样式
    alignment = Alignment(horizontal='center', vertical='center')  # 水平居中、垂直居中

    # 3. 更新数据（从第2行开始，第1行是标题）
    for idx, result_item in enumerate(result_list):
        row_num = idx + 2  # Excel行号从1开始，第1行是标题

        # 写入每个单元格并设置统一的样式
        for col_key, col_num in COL_MAPPING.items():
            cell = ws.cell(row=row_num, column=col_num)
            value = result_item.get(OUTPUT_COLS[col_key])

            # 写入值：简化处理，直接赋值
            if pd.isna(value):
                cell.value = ""
            else:
                cell.value = value

            # 应用统一样式：数字格式+居中对齐
            cell.alignment = alignment

            # 根据列设置不同的数字格式
            if col_key == "rs_col":  # T列：抗力效应比R/S，保留2位小数
                cell.number_format = "0.00"
            else:  # Q、R、S列：保留1位小数
                cell.number_format = "0.0"

    # 4. 确保输出目录存在
    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    # 5. 保存到新文件
    wb.save(save_path)
