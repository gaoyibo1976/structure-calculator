import sys
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 导入核心计算/报告模块
from concrete.core.beam_rect_fc import beam_rect_fc
from concrete.core.report_beam_rect import report_beam_rect_fc
from concrete.core.beam_t_fc import beam_t_fc
from concrete.core.report_beam_t import report_beam_t_fc

# -------------------------- 基础配置 --------------------------
# 数据文件路径
EXCEL_INPUT_PATH = r"/input/梁抗弯承载力数据文件.xlsx"
# 结果输出配置
OUTPUT_DIR = r"/output"
EXCEL_OUTPUT_NAME = "梁抗弯承载力计算结果.xlsx"
EXCEL_OUTPUT_PATH = os.path.join(OUTPUT_DIR, EXCEL_OUTPUT_NAME)
# 抗震承载力调整系数
GAMMA_RE = 0.75
# Excel列定义
OUTPUT_COLS = {
    "x_col": "受压区高度x",  # Q列
    "mu_col": "抗弯承载力Mu",  # R列
    "mue_col": "抗震承载力MuE",  # S列
    "rs_col": "抗力效应比R/S"  # T列
}

# Q-T列的列索引映射
COL_MAPPING = {
    "x_col": 17,  # Q列
    "mu_col": 18,  # R列
    "mue_col": 19,  # S列
    "rs_col": 20  # T列
}


def validate_file_exists(file_path):
    """验证文件是否存在"""
    if not os.path.exists(file_path):
        print(f"❌ 未找到Excel文件！请确认文件路径：{file_path}")
        sys.exit()


def read_excel_data(file_path):
    """读取Excel A-P列数据"""
    try:
        df_input = pd.read_excel(
            file_path,
            sheet_name="Sheet1",
            usecols=["截面编号", "截面类型", "b", "h", "bf", "hf",
                     "混凝土强度等级C", "受拉钢筋强度等级", "受压钢筋强度等级",
                     "受拉钢筋面积As", "受拉钢筋as", "受压钢筋面积As", "受压钢筋as",
                     "弯矩设计值M", "是否地震作用组合", "结构重要性系数γ0"],
            engine="openpyxl",
            dtype={"截面编号": str}
        )
        return df_input
    except Exception as e:
        print(f"❌ 读取Excel文件时出错: {e}")
        sys.exit()


def prepare_calculation_data(df_input):
    """准备计算数据"""
    param = []
    result_data = []

    for _, row in df_input.iterrows():
        # 初始化结果数据
        result_item = row.to_dict()
        for col in OUTPUT_COLS.values():
            result_item[col] = None
        result_data.append(result_item)

        # 构造计算参数
        param_item = {
            "sec_num": row["截面编号"],
            "sec_type": row["截面类型"],
            "M": row["弯矩设计值M"],
            "is_seismic": row["是否地震作用组合"],
            "γ0": row["结构重要性系数γ0"],
            "calc_params": [
                row["b"], row["h"], row["bf"], row["hf"],
                row["混凝土强度等级C"], row["受拉钢筋强度等级"], row["受压钢筋强度等级"],
                row["受拉钢筋面积As"], row["受拉钢筋as"],
                row["受压钢筋面积As"], row["受压钢筋as"], row["结构重要性系数γ0"]
            ]
        }
        param.append(param_item)

    return param, result_data


def calculate_single_item(item, index, total_count):
    """计算单个数据项"""
    sec_num = item["sec_num"] if not pd.isna(item["sec_num"]) else ""
    gamma_0 = item["γ0"]
    sec_num_display = f"序号：{total_count}.{index + 1}      编号：{sec_num}      截面类型：{item['sec_type']}"
    calc_p = item["calc_params"]

    try:
        if item["sec_type"] == "矩形":
            rect_calc_p = calc_p[0:2] + calc_p[4:]  # 跳过bf和hf
            result = beam_rect_fc(*rect_calc_p)
            x = result[0]
            Mu = result[4]
            report = report_beam_rect_fc(sec_num_display, rect_calc_p, result)
            return x, Mu, report, None  # 最后一个参数是错误信息

        elif item["sec_type"] == "T形":
            result = beam_t_fc(*calc_p)
            x = result[1]
            Mu = result[5]
            report = report_beam_t_fc(sec_num_display, calc_p, result)
            return x, Mu, report, None

        else:
            error_msg = f"第{index + 1}行：截面类型'{item['sec_type']}'不支持"
            report = f"【错误】{error_msg}"
            return 0, 0, report, error_msg

    except Exception as e:
        error_msg = f"第{index + 1}行：{str(e)}"
        report = f"【错误】{error_msg}"
        return 0, 0, report, error_msg


def save_excel_result_with_style(result_list, save_path, source_path):
    """保存Excel结果，统一设置样式：数字类型、居中对齐"""
    # 1. 加载原始Excel文件，保留所有样式
    wb = load_workbook(source_path)
    ws = wb.active

    # 2. 定义统一的样式
    alignment = Alignment(horizontal='center', vertical='center')  # 水平居中、垂直居中

    # 3. 更新数据（从第2行开始，第1行是标题）
    for idx, result_item in enumerate(result_list):
        row_num = idx + 2  # Excel行号从1开始，第1行是标题

        # 写入每个单元格并设置统一的样式
        for col_key, col_num in COL_MAPPING.items():
            cell = ws.cell(row=row_num, column=col_num)
            value = result_item.get(OUTPUT_COLS[col_key])

            # 写入值：简化处理，直接赋值
            if pd.isna(value):
                cell.value = ""
            else:
                cell.value = value

            # 应用统一样式：数字格式+居中对齐
            cell.alignment = alignment

            # 根据列设置不同的数字格式
            if col_key == "rs_col":  # T列：抗力效应比R/S，保留2位小数
                cell.number_format = "0.00"
            else:  # Q、R、S列：保留1位小数
                cell.number_format = "0.0"

    # 4. 确保输出目录存在
    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    # 5. 保存到新文件
    wb.save(save_path)


def main():
    """主函数"""
    print("🚀 梁抗弯承载力计算程序启动...")
    start_time = datetime.now()

    # -------------------------- 读取Excel A-P列数据 --------------------------
    print("📖 正在读取Excel文件...")
    validate_file_exists(EXCEL_INPUT_PATH)
    df_input = read_excel_data(EXCEL_INPUT_PATH)

    # -------------------------- 准备计算数据 --------------------------
    param, result_data = prepare_calculation_data(df_input)

    if len(param) == 0:
        print("❌ 未找到有效计算数据，程序终止")
        sys.exit()

    print(f"📊 发现 {len(param)} 组待计算数据")

    # -------------------------- 生成OUT结果文件 --------------------------
    target_dir = OUTPUT_DIR
    os.makedirs(target_dir, exist_ok=True)
    file_name = "梁抗弯承载力计算结果.out"
    file_path = os.path.join(target_dir, file_name)

    local_time = start_time.strftime("%Y-%m-%d %H:%M:%S")

    print("🔄 开始计算...")
    error_count = 0

    with open(file_path, "w", encoding="utf-8") as f:
        f.write(f"{'*' * 52}\n")
        f.write(f"计算时间：{local_time}\n")
        f.write(f"共{len(param)}组截面梁计算数据\n")
        f.write(f"{'*' * 52}\n")

        # 计算每行数据
        for idx, item in enumerate(param):
            # 计算单个数据项
            x, Mu, report, error_msg = calculate_single_item(item, idx, len(param))

            # 记录错误
            if error_msg:
                error_count += 1
                print(f"  ⚠️ {error_msg}")

            # 计算抗震承载力和抗力效应比
            MuE = Mu / GAMMA_RE
            M = item["M"]

            if M == 0 or pd.isna(M):
                R_S = 0
            else:
                is_seismic = item["is_seismic"]
                R_S = MuE / M if is_seismic == 1 else Mu / M

            # 填充Q-T列结果
            result_data[idx][OUTPUT_COLS["x_col"]] = round(x, 3)
            result_data[idx][OUTPUT_COLS["mu_col"]] = round(Mu, 2)
            result_data[idx][OUTPUT_COLS["mue_col"]] = round(MuE, 2)
            result_data[idx][OUTPUT_COLS["rs_col"]] = round(R_S, 2)

            # 写入out文件
            f.write(report + "\n")

        # 写入总结信息
        f.write(f"\n{'=' * 60}\n")
        f.write(f"计算完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"总计: {len(param)} 组数据，其中 {error_count} 组计算出错\n")
        f.write(f"结果文件: {file_path}\n")

    print(f"✅ 计算完成，生成报告文件: {file_path}")
    if error_count > 0:
        print(f"⚠️  注意: 有 {error_count} 组数据计算出错，请查看报告文件")

    # -------------------------- 生成Excel结果文件 --------------------------
    print("💾 正在保存Excel结果...")
    save_excel_result_with_style(result_data, EXCEL_OUTPUT_PATH, EXCEL_INPUT_PATH)
    print("💾 Excel结果文件保存完毕")
    # -------------------------- 程序结束 --------------------------
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()

    print(f"\n🎉 程序执行完毕!")
    print(f"⏱️  总耗时: {duration:.1f}秒")
    print(f"📈 数据处理: {len(param)} 行")
    print(f"📁 输出文件:")
    print(f"   📄 Excel结果: {EXCEL_OUTPUT_PATH}")
    print(f"   📄 详细报告: {file_path}")


if __name__ == "__main__":
    main()