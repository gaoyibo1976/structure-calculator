import sys
import os
import time  # 添加时间模块
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

print(f"🚀 程序启动时间: {datetime.now().strftime('%H:%M:%S')}")
start_total = time.time()

# -------------------------- 基础配置 --------------------------
# 数据文件路径
excel_input_path = r"/input/梁抗弯承载力数据文件.xlsx"
# 结果输出配置
output_dir = r"/output"
excel_output_name = "梁抗弯承载力计算结果.xlsx"
excel_output_path = os.path.join(output_dir, excel_output_name)
# 抗震承载力调整系数
γRE = 0.75
# Excel列定义
output_cols = {
    "x_col": "受压区高度x",  # Q列
    "mu_col": "抗弯承载力Mu",  # R列
    "mue_col": "抗震承载力MuE",  # S列
    "rs_col": "抗力效应比R/S"  # T列
}

# -------------------------- 读取Excel A-P列数据 --------------------------
print("📖 正在读取Excel文件...")
start_read = time.time()

if not os.path.exists(excel_input_path):
    print(f"❌ 未找到Excel文件！请确认文件路径：{excel_input_path}")
    sys.exit()

# 读取A-P列输入数据
df_input = pd.read_excel(
    excel_input_path,
    sheet_name="Sheet1",
    usecols=["截面编号", "截面类型", "b", "h", "bf", "hf",
             "混凝土强度等级C", "受拉钢筋强度等级", "受压钢筋强度等级",
             "受拉钢筋面积As", "受拉钢筋as", "受压钢筋面积As", "受压钢筋as",
             "弯矩设计值M", "是否地震作用组合", "结构重要性系数γ0"],
    engine="openpyxl",
    dtype={"截面编号": str}
)

read_time = time.time() - start_read
print(f"✅ Excel读取完成，耗时: {read_time:.3f}秒，读取到 {len(df_input)} 行数据")

# -------------------------- 构造参数和数据列表 --------------------------
param = []
result_data = []

for index, row in df_input.iterrows():
    result_item = row.to_dict()
    for col in output_cols.values():
        result_item[col] = None
    result_data.append(result_item)

    param_item = {
        "sec_num": row["截面编号"],
        "sec_type": row["截面类型"],
        "M": row["弯矩设计值M"],
        "is_seismic": row["是否地震作用组合"],
        "γ0": row["结构重要性系数γ0"],
        "calc_params": [
            row["b"], row["h"], row["bf"], row["hf"],
            row["混凝土强度等级C"], row["受拉钢筋强度等级"], row["受压钢筋强度等级"],
            row["受拉钢筋面积As"], row["受拉钢筋as"],
            row["受压钢筋面积As"], row["受压钢筋as"], row["结构重要性系数γ0"]
        ]
    }
    param.append(param_item)

# -------------------------- 生成OUT文件 --------------------------
print("🔄 开始计算并生成OUT文件...")
calc_start = time.time()

target_dir = r"/output"
os.makedirs(target_dir, exist_ok=True)
file_name = "梁抗弯承载力计算结果.out"
file_path = os.path.join(target_dir, file_name)

dt = datetime.now()
local_time = dt.strftime("%Y-%m-%d %H:%M:%S")

with open(file_path, "w", encoding="utf-8") as f:
    f.write(f"{'*' * 52}\n")
    f.write(f"计算时间：{local_time}\n")
    f.write(f"共{len(param)}组截面梁计算数据\n")
    f.write(f"{'*' * 52}\n")

    num = 1
    calc_times = []

    for idx, item in enumerate(param):
        start_calc = time.time()

        sec_num = item["sec_num"] if not pd.isna(item["sec_num"]) else ""
        γ0 = item["γ0"]
        sec_num = f"序号：{len(param)}.{num}      编号：{sec_num}      截面类型：{item['sec_type']}"
        calc_p = item["calc_params"]
        report = ""

        # 按截面类型调用计算函数
        if item["sec_type"] == "矩形":
            rect_calc_p = calc_p[0:2] + calc_p[4:]

            # 导入模块（这里会测量导入时间）
            try:
                from concrete.core.beam_rect_fc import beam_rect_fc
                from concrete.core.report_beam_rect import report_beam_rect_fc

                result = beam_rect_fc(*rect_calc_p)
                x = result[0]
                Mu = result[4]
                report = report_beam_rect_fc(sec_num, rect_calc_p, result)
            except Exception as e:
                print(f"❌ 计算第{num}行矩形截面时出错: {e}")
                x = 0
                Mu = 0
                report = f"【错误】序号：{num} 计算出错: {e}"

        elif item["sec_type"] == "T形":
            try:
                from concrete.core.beam_t_fc import beam_t_fc
                from concrete.core.report_beam_t import report_beam_t_fc

                result = beam_t_fc(*calc_p)
                x = result[1]
                Mu = result[5]
                report = report_beam_t_fc(sec_num, calc_p, result)
            except Exception as e:
                print(f"❌ 计算第{num}行T形截面时出错: {e}")
                x = 0
                Mu = 0
                report = f"【错误】序号：{num} 计算出错: {e}"
        else:
            report = f"【错误】序号：{num} 编号：{sec_num} 截面类型{item['sec_type']}不支持！仅支持矩形/T形"
            x = 0
            Mu = 0

        MuE = Mu / γRE
        M = item["M"]

        if M == 0 or pd.isna(M):
            R_S = 0
        else:
            is_seismic = item["is_seismic"]
            R_S = MuE / M if is_seismic == 1 else Mu / M

        # 填充Q-T列结果
        result_data[idx][output_cols["x_col"]] = round(x, 3)
        result_data[idx][output_cols["mu_col"]] = round(Mu, 2)
        result_data[idx][output_cols["mue_col"]] = round(MuE, 2)
        result_data[idx][output_cols["rs_col"]] = round(R_S, 2)

        # 写入out文件
        f.write(report + "\n")

        calc_time = time.time() - start_calc
        calc_times.append(calc_time)

        if calc_time > 0.5:  # 如果单次计算超过0.5秒
            print(f"⚠️ 第{num}行计算耗时较长: {calc_time:.3f}秒")

        num += 1

    # 输出计算统计信息
    if calc_times:
        avg_calc_time = sum(calc_times) / len(calc_times)
        max_calc_time = max(calc_times)
        print(f"📊 计算统计: 平均 {avg_calc_time:.3f}秒/行，最长 {max_calc_time:.3f}秒")

    end_str = f"【END】计算完成，共{len(param)}组数据，结果已保存至：{file_path}"
    f.write(end_str)

calc_total_time = time.time() - calc_start
print(f"✅ 计算完成，总耗时: {calc_total_time:.3f}秒")

# -------------------------- 生成Excel结果文件 --------------------------
print("💾 开始保存Excel文件...")
excel_start = time.time()


def save_excel_result_with_style(result_list, save_path):
    """保存Excel结果，统一设置样式：数字类型、居中对齐"""
    # 1. 加载原始Excel文件，保留所有样式
    wb = load_workbook(excel_input_path)
    ws = wb.active

    # 2. 确定Q-T列的列索引
    col_mapping = {
        "x_col": 17,  # Q列
        "mu_col": 18,  # R列
        "mue_col": 19,  # S列
        "rs_col": 20  # T列
    }

    # 3. 定义统一的样式
    alignment = Alignment(horizontal='center', vertical='center')

    # 4. 更新数据
    for idx, result_item in enumerate(result_list):
        row_num = idx + 2

        for col_key, col_num in col_mapping.items():
            cell = ws.cell(row=row_num, column=col_num)
            value = result_item.get(output_cols[col_key])

            if value is None or pd.isna(value):
                cell.value = ""
            else:
                try:
                    if hasattr(value, 'item'):
                        cell.value = float(value.item())
                    elif isinstance(value, (int, float)):
                        cell.value = float(value)
                    else:
                        cell.value = str(value)
                except Exception:
                    cell.value = ""

            cell.alignment = alignment

            if col_key == "rs_col":
                cell.number_format = "0.00"
            else:
                cell.number_format = "0.0"

    # 5. 保存文件
    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    try:
        wb.save(save_path)
        print(f"✅ Excel结果文件已保存至：{save_path}")
    except Exception as e:
        print(f"❌ 保存Excel文件时出错: {e}")


save_excel_result_with_style(result_data, excel_output_path)

excel_time = time.time() - excel_start
print(f"✅ Excel保存完成，耗时: {excel_time:.3f}秒")

# -------------------------- 程序结束 --------------------------
total_time = time.time() - start_total
print(f"\n🎉 程序执行完毕!")
print(f"⏱️  总运行时间: {total_time:.2f}秒")
print(f"📊 各部分耗时:")
print(f"   - 读取Excel: {read_time:.3f}秒")
print(f"   - 计算: {calc_total_time:.3f}秒")
print(f"   - 保存Excel: {excel_time:.3f}秒")
print(f"   - 其他: {total_time - read_time - calc_total_time - excel_time:.3f}秒")
print(f"\n✅ 计算书文本文件已保存至：{file_path}")