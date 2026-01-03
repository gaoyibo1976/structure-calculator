import sys
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 导入核心计算/报告模块
from concrete.core.beam_rect_fc import beam_rect_fc
from concrete.core.report_beam_rect import report_beam_rect_fc
from concrete.core.beam_t_fc import beam_t_fc
from concrete.core.report_beam_t import report_beam_t_fc

# -------------------------- 基础配置 --------------------------
# 数据文件路径
excel_input_path = r"/input/梁抗弯承载力数据文件.xlsx"
# 结果输出配置
output_dir = r"/output"
excel_output_name = "梁抗弯承载力计算结果.xlsx"
excel_output_path = os.path.join(output_dir, excel_output_name)
# 抗震承载力调整系数
γRE = 0.75
# Excel列定义
output_cols = {
    "x_col": "受压区高度x",  # Q列
    "mu_col": "抗弯承载力Mu",  # R列
    "mue_col": "抗震承载力MuE",  # S列
    "rs_col": "抗力效应比R/S"  # T列
}

# -------------------------- 读取Excel A-P列数据 --------------------------
# 检查文件存在性
if not os.path.exists(excel_input_path):
    print(f"❌ 未找到Excel文件！请确认文件路径：{excel_input_path}")
    sys.exit()

# 读取A-P列输入数据（截面编号为字符型，无需额外转换）
df_input = pd.read_excel(
    excel_input_path,
    sheet_name="Sheet1",
    usecols=["截面编号", "截面类型", "b", "h", "bf", "hf",
             "混凝土强度等级C", "受拉钢筋强度等级", "受压钢筋强度等级",
             "受拉钢筋面积As", "受拉钢筋as", "受压钢筋面积As", "受压钢筋as",
             "弯矩设计值M", "是否地震作用组合", "结构重要性系数γ0"],
    engine="openpyxl",
    dtype={"截面编号": str}  # 显式指定截面编号为字符型（可选，确保读入为字符）
)

# 构造param字典+结果数据列表
param = []
result_data = []

# 核心修改：适配字符型截面编号，仅处理空值，无多余str转换
for index, row in df_input.iterrows():
    # 1. 初始化结果数据（所有行都执行，包括空编号行）
    result_item = row.to_dict()
    for col in output_cols.values():
        result_item[col] = None
    result_data.append(result_item)

    # 2. 构造param字典
    param_item = {
        "sec_num": row["截面编号"],
        "sec_type": row["截面类型"],
        "M": row["弯矩设计值M"],
        "is_seismic": row["是否地震作用组合"],
        "γ0": row["结构重要性系数γ0"],
        "calc_params": [
            row["b"], row["h"], row["bf"], row["hf"],
            row["混凝土强度等级C"], row["受拉钢筋强度等级"], row["受压钢筋强度等级"],
            row["受拉钢筋面积As"], row["受拉钢筋as"],
            row["受压钢筋面积As"], row["受压钢筋as"], row["结构重要性系数γ0"]
        ]
    }
    param.append(param_item)

# 检查有效计算数据（可选：若需完全禁用此检查，可注释/删除）
if len(param) == 0:
    print("❌ Excel文件中无有效计算数据，请检查Sheet1内容！")
    sys.exit()

# -------------------------- 原有out结果文件生成（完全保留） --------------------------
count_total = len(param)  # 所有行数量（含空编号行）

target_dir = r"/output"
os.makedirs(target_dir, exist_ok=True)  # 目标路径如不存在，则创建目录，如已存在，不报错
file_name = "梁抗弯承载力计算结果.out"
file_path = os.path.join(target_dir, file_name)

dt = datetime.now()
local_time = dt.strftime("%Y-%m-%d %H:%M:%S")

with open(file_path, "w", encoding="utf-8") as f:
    f.write(f"{'*' * 52}\n")
    f.write(f"计算时间：{local_time}\n")
    f.write(f"共{count_total}组截面梁计算数据\n")
    f.write(f"{'*' * 52}\n")

    num = 1
    for idx, item in enumerate(param):
        sec_num = item["sec_num"] if not pd.isna(item["sec_num"]) else ""
        γ0 = item["γ0"]
        sec_num = f"序号：{count_total}.{num}      编号：{sec_num}      截面类型：{item['sec_type']}"
        calc_p = item["calc_params"]
        report = ""
        # 按截面类型调用计算函数
        if item["sec_type"] == "矩形":
            rect_calc_p = calc_p[0:2] + calc_p[4:]
            result = beam_rect_fc(*rect_calc_p)
            x = result[0]
            Mu = result[4]
            report = report_beam_rect_fc(sec_num, rect_calc_p, result)

        elif item["sec_type"] == "T形":
            result = beam_t_fc(*calc_p)
            x = result[1]
            Mu = result[5]
            report = report_beam_t_fc(sec_num, calc_p, result)

        else:
            # 空/无效截面类型提示
            report = f"【错误】序号：{num} 编号：{sec_num} 截面类型{item['sec_type']}不支持！仅支持矩形/T形"

        MuE = Mu / γRE
        M = item["M"]

        if M == 0 or pd.isna(M):
            R_S = 0
        else:
            is_seismic = item["is_seismic"]
            R_S = MuE / M if is_seismic == 1 else Mu / M

        # 填充Q-T列结果
        result_data[idx][output_cols["x_col"]] = round(x, 3)
        result_data[idx][output_cols["mu_col"]] = round(Mu, 2)
        result_data[idx][output_cols["mue_col"]] = round(MuE, 2)
        result_data[idx][output_cols["rs_col"]] = round(R_S, 2)

        # 写入out文件并打印
        f.write(report + "\n")

        # 控制台输出，如需恢复取消注释即可
        # print(report)

        num += 1

    end_str = f"【END】计算完成，共{count_total}组数据，结果已保存至：{file_path}"

    # 控制台输出，如需恢复取消注释即可
    # print(end_str)

    f.write(end_str)


# -------------------------- 生成Excel结果文件 --------------------------
def save_excel_result_with_style(result_list, save_path):
    """保存Excel结果，统一设置样式：数字类型、居中对齐"""
    # 1. 加载原始Excel文件，保留所有样式
    wb = load_workbook(excel_input_path)
    ws = wb.active

    # 2. 确定Q-T列的列索引
    col_mapping = {
        "x_col": 17,  # Q列
        "mu_col": 18,  # R列
        "mue_col": 19,  # S列
        "rs_col": 20  # T列
    }

    # 3. 定义统一的样式
    alignment = Alignment(horizontal='center', vertical='center')  # 水平居中、垂直居中

    # 4. 更新数据（从第2行开始，第1行是标题）
    for idx, result_item in enumerate(result_list):
        row_num = idx + 2  # Excel行号从1开始，第1行是标题

        # 写入每个单元格并设置统一的样式
        for col_key, col_num in col_mapping.items():
            cell = ws.cell(row=row_num, column=col_num)

            # 获取值
            value = result_item.get(output_cols[col_key])

            # 写入值
            if value is None or pd.isna(value):
                cell.value = ""
            else:
                # 确保值是Python基本类型
                try:
                    if hasattr(value, 'item'):  # numpy标量
                        cell.value = float(value.item())
                    elif isinstance(value, (int, float)):
                        cell.value = float(value)
                    else:
                        cell.value = str(value)
                except Exception as e:
                    print(f"警告：写入单元格({row_num},{col_num})时出错: {e}，值: {value}")
                    cell.value = ""

            # 应用统一样式：数字格式+居中对齐
            cell.alignment = alignment

            # 根据列设置不同的数字格式
            if col_key == "rs_col":  # T列：抗力效应比R/S，保留2位小数
                cell.number_format = "0.00"
            else:  # Q、R、S列：保留1位小数
                cell.number_format = "0.0"

    # 5. 确保输出目录存在
    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    # 6. 保存到新文件
    try:
        wb.save(save_path)
        print(f"\n✅ Excel结果文件已保存至：{save_path}")
    except Exception as e:
        print(f"❌ 保存Excel文件时出错: {e}")


# 执行保存（使用新函数）
save_excel_result_with_style(result_data, excel_output_path)

# 控制台提示
print(f"\n✅ 计算书文本文件已保存至：{file_path}")