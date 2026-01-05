#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit, QPushButton,
    QVBoxLayout, QHBoxLayout, QGroupBox, QStatusBar, QComboBox, QTextEdit,
    QFileDialog, QMessageBox, QScrollArea, QListWidget, QFrame, QCheckBox
)
from PySide6.QtCore import Qt, QEvent
import sys
import os
import pandas as pd

# 导入计算模块
from concrete.main.梁抗弯承载力计算 import calculate_single_item
from concrete.core.beam_utils import prepare_calculation_data, save_excel_result_with_style

# 添加项目根目录到sys.path
sys.path.append(os.path.abspath(os.path.dirname(__file__)))

class BeamCalculationGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        # 保存当前选中的截面索引
        self.current_section_index = -1
        # 从CHANGELOG.md获取版本号
        self.version = self.get_latest_version()
        self.init_ui()
    
    def get_latest_version(self):
        """从CHANGELOG.md获取最新版本号"""
        try:
            with open("CHANGELOG.md", "r", encoding="utf-8") as f:
                for line in f:
                    if line.startswith("## [") and "Unreleased" not in line:
                        # 提取版本号，格式：## [1.2.0] - 2026-01-04
                        version = line.split("[")[1].split("]")[0]
                        return version
        except Exception as e:
            print(f"读取版本号失败: {e}")
        return "1.0.0"
    
    def init_ui(self):
        """初始化用户界面"""
        # 设置窗口属性，包含版本号
        self.setWindowTitle(f"梁抗弯承载力计算程序 v{self.version}")
        self.setGeometry(100, 100, 1000, 700)
        
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 创建主布局 - 垂直方向
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # 主体内容区域 - 左右分栏
        content_layout = QHBoxLayout()
        
        # 创建左侧单截面参数区
        self.create_left_panel(content_layout)
        
        # 创建右侧批量计算区（不立即加载数据文件）
        self.create_right_panel(content_layout)
        
        main_layout.addLayout(content_layout)
        
        # 底部按钮区域
        self.create_button_panel(main_layout)
        
        # 创建状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        
        # 将状态栏分为两个部分：左侧显示提示信息，右侧显示计算结果
        self.status_bar.showMessage("就绪")
        
        # 创建计算结果显示标签
        self.result_status_label = QLabel(" ")
        self.result_status_label.setAlignment(Qt.AlignRight)
        self.result_status_label.setMinimumWidth(300)  # 设置最小宽度，确保内容完整显示
        
        # 添加到状态栏右侧
        self.status_bar.addPermanentWidget(self.result_status_label)
        
        # 为控件添加悬停提示信息
        self.setup_tooltips()
        
        # 安装事件过滤器
        self.installEventFilters()
        
        # 为参数输入框添加回车键响应
        self.setup_enter_key_response()
        
        # 状态栏创建完成后，再加载数据文件
        self.load_initial_data_file()
    
    def setup_tooltips(self):
        """设置控件的提示信息"""
        # 存储控件和对应的提示信息
        self.tooltip_dict = {
            # 左侧参数面板
            self.sec_num_input: "截面编号，用于标识不同截面",
            self.sec_type_combo: "选择截面类型（矩形或T形）",
            self.b_input: "截面宽度，单位：mm",
            self.h_input: "截面高度，单位：mm",
            self.bf_input: "受压翼缘宽度，单位：mm（仅T形截面可用）",
            self.hf_input: "受压翼缘厚度，单位：mm（仅T形截面可用）",
            self.fcuk_input: "混凝土强度等级，如C30输入30",
            self.fy_combo: "选择受拉钢筋强度等级",
            self.fyc_combo: "选择受压钢筋强度等级",
            self.ast_input: "受拉钢筋面积，单位：mm²",
            self.as_t_input: "受拉钢筋重心到截面受拉边缘的距离，单位：mm",
            self.asc_input: "受压钢筋面积，单位：mm²",
            self.as_c_input: "受压钢筋重心到截面受压边缘的距离，单位：mm",
            self.m_input: "弯矩设计值，单位：kN·m",
            self.seismic_combo: "是否考虑地震作用组合",
            self.seismic_level_combo: "抗震等级，一级到四级",
            self.is_beam_end_combo: "是否为框架梁端",
            self.gamma0_input: "结构重要性系数，一级1.1，二级1.0，三级0.9",
            
            # 右侧批量计算面板
            self.file_input: "数据文件路径",
            self.result_file_input: "结果文件名（无需扩展名，自动生成.xlsx和.out文件）",
            self.output_result_var: "勾选后将生成结果文件",
            self.section_list: "显示数据文件中的截面列表",
            self.result_text: "计算结果输出区域",
            self.save_data_button: "将数据修改保存到数据文件",
            self.add_data_button: "新增截面",
            self.delete_data_button: "删除选中的截面",
            
            # 按钮
            self.calc_button: "进行单个截面抗弯承载力计算",
            self.batch_button: "进行批量截面抗弯承载力计算"
        }
    

    
    def installEventFilters(self):
        """为所有控件安装事件过滤器"""
        for widget in self.tooltip_dict.keys():
            widget.installEventFilter(self)
            
            # 设置鼠标跟踪，以便捕获悬停事件
            widget.setMouseTracking(True)
    
    def eventFilter(self, obj, event):
        """事件过滤器，处理鼠标悬停事件和回车键响应"""
        if event.type() == QEvent.Enter:
            # 鼠标进入控件，显示提示信息（只更新左侧临时状态栏）
            if obj in self.tooltip_dict:
                self.status_bar.showMessage(self.tooltip_dict[obj])
        elif event.type() == QEvent.Leave:
            # 鼠标离开控件，恢复默认状态（只更新左侧临时状态栏）
            self.status_bar.showMessage("就绪")
        elif event.type() == QEvent.KeyPress and event.key() == Qt.Key_Return:
            # 回车键响应：先保存参数，更新列表，再执行计算
            self.save_current_params_to_data()
            self.calculate_single()
        return super().eventFilter(obj, event)
    
    def setup_enter_key_response(self):
        """为参数输入框添加回车键响应，执行单截面计算"""
        # 获取所有参数输入控件
        param_widgets = [
            self.sec_num_input,
            self.sec_type_combo,
            self.b_input,
            self.h_input,
            self.bf_input,
            self.hf_input,
            self.fcuk_input,
            self.fy_combo,
            self.fyc_combo,
            self.ast_input,
            self.as_t_input,
            self.asc_input,
            self.as_c_input,
            self.m_input,
            self.seismic_combo,
            self.seismic_level_combo,
            self.is_beam_end_combo,
            self.gamma0_input
        ]
        
        # 为每个控件添加回车键响应
        for widget in param_widgets:
            # 只为QLineEdit对象连接returnPressed信号
            if hasattr(widget, 'returnPressed'):
                # 回车键响应：先保存参数，更新列表，再执行计算
                widget.returnPressed.connect(self.on_enter_key_pressed)  # 文本框的回车键
            # 为所有控件设置焦点策略，确保能接收焦点，通过事件过滤器处理回车键
            widget.setFocusPolicy(Qt.StrongFocus)  # 确保控件能接收焦点
    
    def on_enter_key_pressed(self):
        """回车键按下时的处理：保存参数、更新列表、执行计算"""
        # 保存当前参数到section_data
        self.save_current_params_to_data()
        # 执行单截面计算
        self.calculate_single()
    
    def create_left_panel(self, parent_layout):
        """创建左侧参数面板"""
        left_group = QGroupBox("单个截面参数")
        left_layout = QVBoxLayout(left_group)
        
        # 设置左侧面板的固定宽度，窗口最大化时保持不变
        left_group.setFixedWidth(300)
        
        # 截面编号
        sec_num_layout = QHBoxLayout()
        sec_num_label = QLabel("截面编号:")
        self.sec_num_input = QLineEdit("")
        self.sec_num_input.setFixedWidth(150)
        sec_num_layout.addWidget(sec_num_label)
        sec_num_layout.addWidget(self.sec_num_input)
        left_layout.addLayout(sec_num_layout)
        
        # 截面类型
        sec_type_layout = QHBoxLayout()
        sec_type_label = QLabel("截面类型:")
        self.sec_type_combo = QComboBox()
        self.sec_type_combo.addItems(["矩形", "T形"])
        self.sec_type_combo.setFixedWidth(100)
        self.sec_type_combo.currentTextChanged.connect(self.on_section_type_changed)
        sec_type_layout.addWidget(sec_type_label)
        sec_type_layout.addWidget(self.sec_type_combo)
        left_layout.addLayout(sec_type_layout)
        
        # 梁宽 b
        b_layout = QHBoxLayout()
        b_label = QLabel("截面 b (mm):")
        self.b_input = QLineEdit("300")
        self.b_input.setFixedWidth(100)
        b_layout.addWidget(b_label)
        b_layout.addWidget(self.b_input)
        left_layout.addLayout(b_layout)
        
        # 梁高 h
        h_layout = QHBoxLayout()
        h_label = QLabel("梁高 h (mm):")
        self.h_input = QLineEdit("600")
        self.h_input.setFixedWidth(100)
        h_layout.addWidget(h_label)
        h_layout.addWidget(self.h_input)
        left_layout.addLayout(h_layout)
        
        # 受压翼缘宽度 bf'
        bf_layout = QHBoxLayout()
        self.bf_label = QLabel("受压翼缘宽度 bf' (mm):")
        self.bf_input = QLineEdit("0")
        self.bf_input.setFixedWidth(100)
        bf_layout.addWidget(self.bf_label)
        bf_layout.addWidget(self.bf_input)
        left_layout.addLayout(bf_layout)
        
        # 受压翼缘厚度 hf'
        hf_layout = QHBoxLayout()
        self.hf_label = QLabel("受压翼缘厚度 hf' (mm):")
        self.hf_input = QLineEdit("0")
        self.hf_input.setFixedWidth(100)
        hf_layout.addWidget(self.hf_label)
        hf_layout.addWidget(self.hf_input)
        left_layout.addLayout(hf_layout)
        
        # 混凝土强度等级
        fcuk_layout = QHBoxLayout()
        fcuk_label = QLabel("混凝土强度等级 C:")
        self.fcuk_input = QLineEdit("30")
        self.fcuk_input.setFixedWidth(100)
        fcuk_layout.addWidget(fcuk_label)
        fcuk_layout.addWidget(self.fcuk_input)
        left_layout.addLayout(fcuk_layout)
        
        # 受拉钢筋强度等级
        fy_layout = QHBoxLayout()
        fy_label = QLabel("受拉钢筋强度等级:")
        self.fy_combo = QComboBox()
        self.fy_combo.addItems(["HPB300", "HRB335", "HRB400", "HRB500"])
        self.fy_combo.setCurrentText("HRB400")
        self.fy_combo.setFixedWidth(100)
        fy_layout.addWidget(fy_label)
        fy_layout.addWidget(self.fy_combo)
        left_layout.addLayout(fy_layout)
        
        # 受压钢筋强度等级
        fyc_layout = QHBoxLayout()
        fyc_label = QLabel("受压钢筋强度等级:")
        self.fyc_combo = QComboBox()
        self.fyc_combo.addItems(["HPB300", "HRB335", "HRB400", "HRB500"])
        self.fyc_combo.setCurrentText("HRB400")
        self.fyc_combo.setFixedWidth(100)
        fyc_layout.addWidget(fyc_label)
        fyc_layout.addWidget(self.fyc_combo)
        left_layout.addLayout(fyc_layout)
        
        # 受拉钢筋面积
        ast_layout = QHBoxLayout()
        ast_label = QLabel("受拉钢筋面积 As (mm²):")
        self.ast_input = QLineEdit("1500")
        self.ast_input.setFixedWidth(100)
        ast_layout.addWidget(ast_label)
        ast_layout.addWidget(self.ast_input)
        left_layout.addLayout(ast_layout)
        
        # 受拉钢筋as
        as_t_layout = QHBoxLayout()
        as_t_label = QLabel("受拉钢筋 as (mm):")
        self.as_t_input = QLineEdit("42.5")
        self.as_t_input.setFixedWidth(100)
        as_t_layout.addWidget(as_t_label)
        as_t_layout.addWidget(self.as_t_input)
        left_layout.addLayout(as_t_layout)
        
        # 受压钢筋面积
        asc_layout = QHBoxLayout()
        asc_label = QLabel("受压钢筋面积 As' (mm²):")
        self.asc_input = QLineEdit("0")
        self.asc_input.setFixedWidth(100)
        asc_layout.addWidget(asc_label)
        asc_layout.addWidget(self.asc_input)
        left_layout.addLayout(asc_layout)
        
        # 受压钢筋as'
        as_c_layout = QHBoxLayout()
        as_c_label = QLabel("受压钢筋 as' (mm):")
        self.as_c_input = QLineEdit("42.5")
        self.as_c_input.setFixedWidth(100)
        as_c_layout.addWidget(as_c_label)
        as_c_layout.addWidget(self.as_c_input)
        left_layout.addLayout(as_c_layout)
        
        # 弯矩设计值 M
        m_layout = QHBoxLayout()
        m_label = QLabel("弯矩设计值 M (kN·m):")
        self.m_input = QLineEdit("250")
        self.m_input.setFixedWidth(100)
        m_layout.addWidget(m_label)
        m_layout.addWidget(self.m_input)
        left_layout.addLayout(m_layout)
        
        # 是否地震作用组合
        seismic_layout = QHBoxLayout()
        seismic_label = QLabel("是否地震作用组合:")
        self.seismic_combo = QComboBox()
        self.seismic_combo.addItems(["否", "是"])
        self.seismic_combo.setFixedWidth(100)
        seismic_layout.addWidget(seismic_label)
        seismic_layout.addWidget(self.seismic_combo)
        left_layout.addLayout(seismic_layout)
        
        # 抗震等级
        seismic_level_layout = QHBoxLayout()
        seismic_level_label = QLabel("抗震等级:")
        self.seismic_level_combo = QComboBox()
        self.seismic_level_combo.addItems(["一级", "二级", "三级", "四级"])
        self.seismic_level_combo.setFixedWidth(100)
        seismic_level_layout.addWidget(seismic_level_label)
        seismic_level_layout.addWidget(self.seismic_level_combo)
        left_layout.addLayout(seismic_level_layout)
        
        # 是否框架梁端
        is_beam_end_layout = QHBoxLayout()
        is_beam_end_label = QLabel("是否框架梁端:")
        self.is_beam_end_combo = QComboBox()
        self.is_beam_end_combo.addItems(["否", "是"])
        self.is_beam_end_combo.setFixedWidth(100)
        is_beam_end_layout.addWidget(is_beam_end_label)
        is_beam_end_layout.addWidget(self.is_beam_end_combo)
        left_layout.addLayout(is_beam_end_layout)
        
        # 结构重要性系数 γ0
        gamma0_layout = QHBoxLayout()
        gamma0_label = QLabel("结构重要性系数 γ0:")
        self.gamma0_input = QLineEdit("1.0")
        self.gamma0_input.setFixedWidth(100)
        gamma0_layout.addWidget(gamma0_label)
        gamma0_layout.addWidget(self.gamma0_input)
        left_layout.addLayout(gamma0_layout)
        
        # 初始设置
        self.on_section_type_changed("矩形")
        
        parent_layout.addWidget(left_group, 1)
    
    def on_section_type_changed(self, sec_type):
        """截面类型变化时的处理"""
        if sec_type == "矩形":
            # 矩形截面，禁用受压翼缘参数
            self.bf_input.setEnabled(False)
            self.hf_input.setEnabled(False)
            self.bf_label.setEnabled(False)
            self.hf_label.setEnabled(False)
        else:
            # T形截面，启用受压翼缘参数
            self.bf_input.setEnabled(True)
            self.hf_input.setEnabled(True)
            self.bf_label.setEnabled(True)
            self.hf_label.setEnabled(True)
    
    def create_right_panel(self, parent_layout):
        """创建右侧批量计算面板"""
        right_group = QGroupBox("批量计算")
        right_layout = QVBoxLayout(right_group)
        
        # 默认工作目录：D:\Work
        self.default_work_dir = "D:\\Work"
        if not os.path.exists(self.default_work_dir):
            os.makedirs(self.default_work_dir, exist_ok=True)
        
        # 数据文件选择
        file_layout = QHBoxLayout()
        file_label = QLabel("数据文件:")
        default_data_file = os.path.join(self.default_work_dir, "梁抗弯承载力数据文件.xlsx")
        self.file_input = QLineEdit(default_data_file)
        browse_button = QPushButton("浏览...")
        browse_button.setFixedWidth(80)
        browse_button.clicked.connect(self.browse_file)
        
        file_layout.addWidget(file_label)
        file_layout.addWidget(self.file_input)
        file_layout.addWidget(browse_button)
        right_layout.addLayout(file_layout)
        
        # 结果文件设置 - 只需要输入文件名
        result_file_layout = QHBoxLayout()
        result_file_label = QLabel("结果文件名:")
        self.result_file_input = QLineEdit("抗弯承载力计算结果")
        result_file_layout.addWidget(result_file_label)
        result_file_layout.addWidget(self.result_file_input)
        
        # 添加.xlsx及.out的文字提示
        file_ext_label = QLabel(".xlsx及.out")
        result_file_layout.addWidget(file_ext_label)
        
        # 添加将数据修改保存到数据文件按钮
        self.save_data_button = QPushButton("保存数据到文件")
        self.save_data_button.setFixedWidth(120)
        self.save_data_button.clicked.connect(self.save_data_to_file)
        result_file_layout.addWidget(self.save_data_button)
        
        # 添加新增截面按钮
        self.add_data_button = QPushButton("新增截面")
        self.add_data_button.setFixedWidth(80)
        self.add_data_button.clicked.connect(self.add_new_section)
        result_file_layout.addWidget(self.add_data_button)
        
        # 添加删除截面按钮
        self.delete_data_button = QPushButton("删除截面")
        self.delete_data_button.setFixedWidth(80)
        self.delete_data_button.clicked.connect(self.delete_selected_section)
        result_file_layout.addWidget(self.delete_data_button)
        right_layout.addLayout(result_file_layout)
        
        # 输出结果文件复选框
        output_check_layout = QHBoxLayout()
        self.output_result_var = QCheckBox("输出结果文件")
        self.output_result_var.setChecked(True)  # 默认勾选
        output_check_layout.addWidget(self.output_result_var)
        right_layout.addLayout(output_check_layout)
        
        # 主体内容区域 - 列表框和文本输出区
        main_content_layout = QHBoxLayout()
        
        # 左侧列表框（宽度减小一半）
        list_frame = QFrame()
        list_frame.setFrameStyle(QFrame.StyledPanel)
        list_layout = QVBoxLayout(list_frame)
        list_layout.setContentsMargins(5, 5, 5, 5)
        
        # 设置列表框的固定宽度，窗口最大化时保持不变
        list_frame.setFixedWidth(150)
        
        # 列表框
        self.section_list = QListWidget()
        self.section_list.itemClicked.connect(self.on_list_item_clicked)
        list_layout.addWidget(QLabel("截面列表"))
        list_layout.addWidget(self.section_list)
        main_content_layout.addWidget(list_frame)  # 固定宽度，不需要比例
        
        # 右侧文本输出区（相应加宽）
        text_frame = QFrame()
        text_frame.setFrameStyle(QFrame.StyledPanel)
        text_layout = QVBoxLayout(text_frame)
        text_layout.setContentsMargins(5, 5, 5, 5)
        
        # 文本输出
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        self.result_text.setFontFamily("Consolas")
        self.result_text.setFontPointSize(10)
        text_layout.addWidget(QLabel("文本输出"))
        text_layout.addWidget(self.result_text)
        
        main_content_layout.addWidget(text_frame, 3)  # 宽度比例3，从2调整为3
        
        right_layout.addLayout(main_content_layout)
        
        parent_layout.addWidget(right_group, 2)
        
        # 存储数据文件中的截面数据
        self.section_data = []
        # 初始数据文件路径
        self.default_data_file = default_data_file
    
    def browse_file(self):
        """浏览文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择数据文件", self.default_work_dir, "Excel 文件 (*.xlsx);;所有文件 (*.*)"
        )
        if file_path:
            self.file_input.setText(file_path)
            # 文件选择后，自动读取数据文件并更新列表框
            self.load_data_file_to_list(file_path)
    
    def load_data_file_to_list(self, file_path):
        """加载数据文件到列表框"""
        try:
            self.status_bar.showMessage(f"正在读取数据文件: {os.path.basename(file_path)}")
            
            # 读取Excel文件
            df = pd.read_excel(file_path)
            
            # 清空列表和数据存储
            self.section_list.clear()
            self.section_data.clear()
            
            # 遍历数据行，添加到列表框
            for idx, row in df.iterrows():
                # 构造列表项文本，格式为：序号-编号，使用数据文件中的"截面编号"字段
                sec_num = str(row.get("截面编号", row.get("sec_num", f"截面{idx+1}")))
                list_text = f"{idx+1}-{sec_num}"
                self.section_list.addItem(list_text)
                
                # 处理数据，将nan值替换为0
                row_dict = row.to_dict()
                # 遍历所有值，将nan替换为0
                for key, value in row_dict.items():
                    if pd.isna(value):
                        row_dict[key] = 0
                # 存储处理后的数据
                self.section_data.append(row_dict)
            
            self.status_bar.showMessage(f"已读取 {len(self.section_data)} 个截面数据")
            
            # 默认选择第一个截面
            if self.section_data:
                self.on_list_item_clicked(self.section_list.item(0))
        
        except Exception as e:
            error_msg = f"读取数据文件失败: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
    
    def on_list_item_clicked(self, item):
        """列表项点击事件处理"""
        if not item:
            return
        
        try:
            # 保存当前截面的数据（如果有选中的截面）
            if self.current_section_index >= 0:
                self.save_current_params_to_data()
            
            # 获取点击的列表项索引
            index = self.section_list.row(item)
            if 0 <= index < len(self.section_data):
                # 保存当前选中的截面索引
                self.current_section_index = index
                
                # 获取对应的数据
                data = self.section_data[index]
                
                # 更新左侧参数面板
                self.update_parameter_panel(data)
                
                # 状态栏提示
                self.status_bar.showMessage(f"已加载截面: {data.get('截面编号', data.get('sec_num', f'截面{index+1}'))}")
                
                # 自动执行计算
                self.calculate_single()
        
        except Exception as e:
            error_msg = f"加载截面数据失败: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
    
    def update_parameter_panel(self, data):
        """根据数据更新参数面板"""
        # 调试信息：打印数据内容，查看实际字段名
        print(f"更新参数面板，数据: {data}")
        
        # 更新截面编号
        sec_num = data.get("截面编号", data.get("sec_num", ""))
        self.sec_num_input.setText(str(sec_num))
        
        # 更新截面类型
        # 尝试多种可能的字段名
        sec_type = data.get("sec_type", data.get("截面类型", "矩形"))
        self.sec_type_combo.setCurrentText(sec_type)
        
        # 手动调用截面类型变化处理函数，确保截面类型状态正确
        self.on_section_type_changed(sec_type)
        
        # 更新截面尺寸
        self.b_input.setText(str(data.get("b", data.get("截面宽度", 300))))
        self.h_input.setText(str(data.get("h", data.get("截面高度", 600))))
        self.bf_input.setText(str(data.get("bf", data.get("受压翼缘宽度", 0))))
        self.hf_input.setText(str(data.get("hf", data.get("受压翼缘厚度", 0))))
        
        # 更新混凝土强度等级
        self.fcuk_input.setText(str(data.get("fcuk", data.get("混凝土强度等级", 30))))
        
        # 更新钢筋强度等级
        fy_grade = data.get("fy_grade", data.get("受拉钢筋强度等级", data.get("钢筋强度等级", "HRB400")))
        self.fy_combo.setCurrentText(fy_grade)
        
        # 更新受压钢筋强度等级
        fyc_grade = data.get("fyc_grade", data.get("受压钢筋强度等级", fy_grade))
        self.fyc_combo.setCurrentText(fyc_grade)
        
        # 更新钢筋面积和位置
        # 尝试多种可能的字段名，确保能获取到正确的数据
        ast = data.get("ast", data.get("As", data.get("受拉钢筋面积", data.get("受拉钢筋面积As", 1500))))
        asc = data.get("asc", data.get("As'", data.get("受压钢筋面积", data.get("受压钢筋面积As", 0))))
        self.ast_input.setText(str(ast))
        self.as_t_input.setText(str(data.get("as_t", data.get("as", data.get("受拉钢筋as", 42.5)))))
        self.asc_input.setText(str(asc))
        self.as_c_input.setText(str(data.get("as_c", data.get("as'", data.get("受压钢筋as", data.get("受压钢筋as'", 42.5))))))
        
        # 更新弯矩设计值和其他参数
        self.m_input.setText(str(data.get("M", data.get("弯矩设计值", data.get("弯矩", 250)))))
        
        # 更新地震作用组合
        is_seismic = data.get("is_seismic", data.get("是否地震", data.get("地震作用组合", data.get("是否地震作用组合", 0))))
        seismic_text = "是" if is_seismic == 1 else "否"
        self.seismic_combo.setCurrentText(seismic_text)
        
        # 更新抗震等级
        seismic_level = data.get("seismic_level", data.get("抗震等级", "二级"))
        self.seismic_level_combo.setCurrentText(seismic_level)
        
        # 更新是否框架梁端
        is_beam_end = data.get("is_beam_end", data.get("是否框架梁端", 0))
        beam_end_text = "是" if is_beam_end == 1 else "否"
        self.is_beam_end_combo.setCurrentText(beam_end_text)
        
        # 更新结构重要性系数
        self.gamma0_input.setText(str(data.get("γ0", data.get("gamma0", data.get("结构重要性系数", 1.0)))))
        
        # 确保所有控件都更新完成
        self.status_bar.showMessage(f"已更新截面参数: {data.get('sec_num', '未知')}")    
    
    def save_current_params_to_data(self):
        """将当前参数面板的数据保存到section_data中"""
        if self.current_section_index >= 0 and self.current_section_index < len(self.section_data):
            try:
                # 获取当前参数面板的数据
                sec_num = self.sec_num_input.text().strip()
                sec_type = self.sec_type_combo.currentText()
                b = float(self.b_input.text())
                h = float(self.h_input.text())
                bf = float(self.bf_input.text())
                hf = float(self.hf_input.text())
                fcuk = float(self.fcuk_input.text())
                fy_grade = self.fy_combo.currentText()
                fyc_grade = self.fyc_combo.currentText()
                ast = float(self.ast_input.text())
                as_t = float(self.as_t_input.text())
                asc = float(self.asc_input.text())
                as_c = float(self.as_c_input.text())
                M = float(self.m_input.text())
                is_seismic = 1 if self.seismic_combo.currentText() == "是" else 0
                seismic_level = self.seismic_level_combo.currentText()
                is_beam_end = 1 if self.is_beam_end_combo.currentText() == "是" else 0
                gamma0 = float(self.gamma0_input.text())
                
                # 如果截面编号为空，使用默认值
                if not sec_num:
                    sec_num = f"截面{self.current_section_index+1}"
                
                # 获取原始数据的所有列，确保只更新读取的单元格
                data_dict = self.section_data[self.current_section_index].copy()
                
                # 更新数据字典，只更新已读取的字段
                data_dict.update({
                    "截面编号": sec_num,
                    "sec_num": sec_num,
                    "γ0": gamma0,
                    "gamma0": gamma0,
                    "结构重要性系数": gamma0,
                    "M": M,
                    "弯矩设计值": M,
                    "弯矩": M,
                    "is_seismic": is_seismic,
                    "是否地震": is_seismic,
                    "地震作用组合": is_seismic,
                    "是否地震作用组合": is_seismic,
                    "抗震等级": seismic_level,
                    "seismic_level": seismic_level,
                    "是否框架梁端": is_beam_end,
                    "is_beam_end": is_beam_end,
                    "sec_type": sec_type,
                    "截面类型": sec_type,
                    "b": b,
                    "截面宽度": b,
                    "h": h,
                    "截面高度": h,
                    "bf": bf,
                    "受压翼缘宽度": bf,
                    "hf": hf,
                    "受压翼缘厚度": hf,
                    "fcuk": fcuk,
                    "混凝土强度等级": fcuk,
                    "fy_grade": fy_grade,
                    "受拉钢筋强度等级": fy_grade,
                    "钢筋强度等级": fy_grade,
                    "fyc_grade": fyc_grade,
                    "受压钢筋强度等级": fyc_grade,
                    "ast": ast,
                    "As": ast,
                    "受拉钢筋面积": ast,
                    "受拉钢筋面积As": ast,
                    "as_t": as_t,
                    "as": as_t,
                    "受拉钢筋as": as_t,
                    "asc": asc,
                    "As'": asc,
                    "受压钢筋面积": asc,
                    "受压钢筋面积As": asc,
                    "as_c": as_c,
                    "as'": as_c,
                    "受压钢筋as": as_c,
                    "受压钢筋as'": as_c
                })
                
                # 更新section_data
                self.section_data[self.current_section_index] = data_dict
                
                # 更新列表框中的显示文本
                if self.section_list.count() > self.current_section_index:
                    list_item = self.section_list.item(self.current_section_index)
                    if list_item:
                        new_text = f"{self.current_section_index+1}-{sec_num}"
                        list_item.setText(new_text)
                
                print(f"已保存当前截面数据到section_data: {data_dict}")
                
            except Exception as e:
                print(f"保存当前参数到section_data失败: {e}")
    
    def load_initial_data_file(self):
        """加载初始数据文件"""
        # 初始加载默认数据文件到列表框
        self.load_data_file_to_list(self.default_data_file)
        
    def add_new_section(self):
        """新增截面，自动增加一个列表，并赋予默认参数"""
        try:
            # 先保存当前参数到section_data
            self.save_current_params_to_data()
            
            self.status_bar.showMessage("正在新增数据...")
            
            # 生成新的截面编号，序号自动加1
            new_idx = len(self.section_data)
            new_sec_num = f"截面{new_idx+1}"
            
            # 如果已有数据，使用最后一行数据作为模板，确保所有列都存在
            if self.section_data:
                # 复制最后一行数据作为模板
                default_data = self.section_data[-1].copy()
                # 更新关键参数
                default_data.update({
                    "截面编号": new_sec_num,
                    "sec_num": new_sec_num,
                    "γ0": 1.0,
                    "gamma0": 1.0,
                    "结构重要性系数": 1.0,
                    "M": 250,
                    "弯矩设计值": 250,
                    "弯矩": 250,
                    "is_seismic": 0,
                    "是否地震": 0,
                    "地震作用组合": 0,
                    "是否地震作用组合": 0,
                    "sec_type": "矩形",
                    "截面类型": "矩形",
                    "b": 300,
                    "截面宽度": 300,
                    "h": 600,
                    "截面高度": 600,
                    "bf": 0,
                    "受压翼缘宽度": 0,
                    "hf": 0,
                    "受压翼缘厚度": 0,
                    "fcuk": 30,
                    "混凝土强度等级": 30,
                    "fy_grade": "HRB400",
                    "受拉钢筋强度等级": "HRB400",
                    "钢筋强度等级": "HRB400",
                    "fyc_grade": "HRB400",
                    "受压钢筋强度等级": "HRB400",
                    "ast": 1500,
                    "As": 1500,
                    "受拉钢筋面积": 1500,
                    "受拉钢筋面积As": 1500,
                    "as_t": 42.5,
                    "as": 42.5,
                    "受拉钢筋as": 42.5,
                    "asc": 0,
                    "As'": 0,
                    "受压钢筋面积": 0,
                    "受压钢筋面积As": 0,
                    "as_c": 42.5,
                    "as'": 42.5,
                    "受压钢筋as": 42.5,
                    "受压钢筋as'": 42.5
                })
            else:
                # 如果没有数据，使用默认参数
                default_data = {
                    "截面编号": new_sec_num,
                    "sec_num": new_sec_num,
                    "γ0": 1.0,
                    "gamma0": 1.0,
                    "结构重要性系数": 1.0,
                    "M": 250,
                    "弯矩设计值": 250,
                    "弯矩": 250,
                    "is_seismic": 0,
                    "是否地震": 0,
                    "地震作用组合": 0,
                    "是否地震作用组合": 0,
                    "sec_type": "矩形",
                    "截面类型": "矩形",
                    "b": 300,
                    "截面宽度": 300,
                    "h": 600,
                    "截面高度": 600,
                    "bf": 0,
                    "受压翼缘宽度": 0,
                    "hf": 0,
                    "受压翼缘厚度": 0,
                    "fcuk": 30,
                    "混凝土强度等级": 30,
                    "fy_grade": "HRB400",
                    "受拉钢筋强度等级": "HRB400",
                    "钢筋强度等级": "HRB400",
                    "fyc_grade": "HRB400",
                    "受压钢筋强度等级": "HRB400",
                    "ast": 1500,
                    "As": 1500,
                    "受拉钢筋面积": 1500,
                    "受拉钢筋面积As": 1500,
                    "as_t": 42.5,
                    "as": 42.5,
                    "受拉钢筋as": 42.5,
                    "asc": 0,
                    "As'": 0,
                    "受压钢筋面积": 0,
                    "受压钢筋面积As": 0,
                    "as_c": 42.5,
                    "as'": 42.5,
                    "受压钢筋as": 42.5,
                    "受压钢筋as'": 42.5
                }
            
            # 添加到section_data列表
            self.section_data.append(default_data)
            
            # 更新截面列表
            list_text = f"{new_idx+1}-{new_sec_num}"
            self.section_list.addItem(list_text)
            
            # 自动选择新添加的数据
            self.section_list.setCurrentRow(new_idx)
            self.on_list_item_clicked(self.section_list.item(new_idx))
            
            self.status_bar.showMessage(f"已成功新增数据，序号: {new_idx+1}")
            self.result_text.append(f"已成功新增数据，序号: {new_idx+1}\n")
            
        except Exception as e:
            error_msg = f"新增数据失败: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
    
    def delete_selected_section(self):
        """删除选中的截面"""
        try:
            # 获取当前选中的列表项
            current_item = self.section_list.currentItem()
            if not current_item:
                self.status_bar.showMessage("请先选择要删除的截面数据")
                return
            
            # 获取选中项的索引
            index = self.section_list.row(current_item)
            if 0 <= index < len(self.section_data):
                # 获取要删除的截面编号
                sec_num = self.section_data[index].get("sec_num", self.section_data[index].get("截面编号", f"截面{index+1}"))
                
                # 从数据存储和列表框中删除
                del self.section_data[index]
                self.section_list.takeItem(index)
                
                # 更新列表项的显示文本（重新编号）
                for i in range(self.section_list.count()):
                    item = self.section_list.item(i)
                    new_sec_num = self.section_data[i].get("sec_num", self.section_data[i].get("截面编号", f"截面{i+1}"))
                    item.setText(f"{i+1}-{new_sec_num}")
                
                # 如果还有数据，自动选中第一个截面
                if self.section_data:
                    self.section_list.setCurrentRow(0)
                    self.on_list_item_clicked(self.section_list.currentItem())
                
                self.status_bar.showMessage(f"已删除截面数据: {sec_num}")
                self.result_text.append(f"已删除截面数据: {sec_num}\n")
        
        except Exception as e:
            error_msg = f"删除数据失败: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
    
    def save_data_to_file(self):
        """将数据修改保存到数据文件，只回写A-P列（原始数据），保持原文件样式"""
        try:
            # 先保存当前参数到section_data
            self.save_current_params_to_data()
            
            self.status_bar.showMessage("正在保存数据到文件...")
            
            # 获取文件路径
            file_path = self.file_input.text()
            
            # 使用openpyxl加载文件，保持样式
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 定义允许修改的列范围：A-P列（Excel列索引1-16，因为Excel从1开始计数）
            allowed_cols = list(range(1, 17))  # A-P列对应1-16
            
            # 定义列索引到数据的映射关系（A-P列）
            col_data_mapping = {
                1: "截面编号",        # A列
                2: "结构重要性系数γ0",  # B列
                3: "弯矩设计值M",     # C列
                4: "是否地震作用组合",    # D列
                5: "截面类型",        # E列
                6: "b",             # F列
                7: "h",             # G列
                8: "bf",            # H列
                9: "hf",            # I列
                10: "混凝土强度等级C",    # J列
                11: "受拉钢筋强度等级",   # K列
                12: "受压钢筋强度等级",   # L列
                13: "受拉钢筋面积As",    # M列
                14: "受拉钢筋as",     # N列
                15: "受压钢筋面积As",    # O列
                16: "受压钢筋as"      # P列
            }
            
            # 处理每一行数据
            for idx, data in enumerate(self.section_data):
                # 转换为Excel行号（从2开始，因为表头占1行）
                excel_row = idx + 2
                
                # 确保行存在，如果不存在则添加，使用insert_rows保持样式
                while excel_row > ws.max_row:
                    # 复制前一行的样式到新行
                    ws.insert_rows(ws.max_row + 1)
                    # 清空新行的A-P列内容，但保留样式
                    for col_idx in allowed_cols:
                        ws.cell(row=excel_row, column=col_idx, value="")
                
                # 只更新A-P列
                for col_idx in allowed_cols:
                    # 获取对应的数据键
                    data_key = col_data_mapping[col_idx]
                    
                    # 根据数据键获取对应的值
                    cell_value = data.get(data_key, "")
                    
                    # 特殊处理一些字段
                    if data_key == "结构重要性系数γ0":
                        cell_value = data.get("γ0", data.get("gamma0", data.get("结构重要性系数", 1.0)))
                    elif data_key == "弯矩设计值M":
                        cell_value = data.get("M", data.get("弯矩设计值", data.get("弯矩", 250)))
                    elif data_key == "是否地震作用组合":
                        cell_value = data.get("is_seismic", data.get("是否地震", data.get("地震作用组合", data.get("是否地震作用组合", 0))))
                    elif data_key == "截面类型":
                        cell_value = data.get("sec_type", data.get("截面类型", "矩形"))
                    elif data_key == "b":
                        cell_value = data.get("b", data.get("截面宽度", 300))
                    elif data_key == "h":
                        cell_value = data.get("h", data.get("截面高度", 600))
                    elif data_key == "bf":
                        cell_value = data.get("bf", data.get("受压翼缘宽度", 0))
                    elif data_key == "hf":
                        cell_value = data.get("hf", data.get("受压翼缘厚度", 0))
                    elif data_key == "混凝土强度等级C":
                        cell_value = data.get("fcuk", data.get("混凝土强度等级", 30))
                    elif data_key == "受拉钢筋强度等级":
                        cell_value = data.get("fy_grade", data.get("受拉钢筋强度等级", data.get("钢筋强度等级", "HRB400")))
                    elif data_key == "受压钢筋强度等级":
                        cell_value = data.get("fyc_grade", data.get("受压钢筋强度等级", "HRB400"))
                    elif data_key == "受拉钢筋面积As":
                        cell_value = data.get("ast", data.get("As", data.get("受拉钢筋面积", data.get("受拉钢筋面积As", 1500))))
                    elif data_key == "受拉钢筋as":
                        cell_value = data.get("as_t", data.get("as", data.get("受拉钢筋as", 42.5)))
                    elif data_key == "受压钢筋面积As":
                        cell_value = data.get("asc", data.get("As'", data.get("受压钢筋面积", data.get("受压钢筋面积As", 0))))
                    elif data_key == "受压钢筋as":
                        cell_value = data.get("as_c", data.get("as'", data.get("受压钢筋as", data.get("受压钢筋as'", 42.5))))
                    elif data_key == "截面编号":
                        cell_value = data.get("截面编号", data.get("sec_num", f"截面{idx+1}"))
                    
                    # 设置单元格值，openpyxl会保持原有样式
                    ws.cell(row=excel_row, column=col_idx, value=cell_value)
            
            # 保存文件，保持样式
            wb.save(file_path)
            
            # 显示保存成功信息
            self.status_bar.showMessage(f"数据已成功保存到文件: {file_path}")
            self.result_text.append(f"数据已成功保存到文件: {file_path}\n")
            
        except Exception as e:
            error_msg = f"保存数据到文件失败: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
    
    def create_button_panel(self, parent_layout):
        """创建底部按钮面板"""
        button_frame = QFrame()
        button_layout = QHBoxLayout(button_frame)
        button_layout.setContentsMargins(0, 0, 0, 0)
        
        # 单截面计算按钮 - 宽度与参数区匹配
        self.calc_button = QPushButton("单截面计算")
        self.calc_button.setFixedHeight(40)
        self.calc_button.setFixedWidth(300)  # 与左侧参数面板宽度相同
        self.calc_button.clicked.connect(self.calculate_single)
        button_layout.addWidget(self.calc_button)
        
        # 批量计算按钮 - 宽度与批量计算区匹配，占据剩余空间
        self.batch_button = QPushButton("批量计算")
        self.batch_button.setFixedHeight(40)
        self.batch_button.setMinimumHeight(40)
        self.batch_button.clicked.connect(self.calculate_batch)
        button_layout.addWidget(self.batch_button)
        
        parent_layout.addWidget(button_frame)
    
    def calculate_single(self):
        """单个截面计算"""
        try:
            # 清空文本输出框
            self.result_text.clear()
            
            # 状态栏提示
            self.status_bar.showMessage("正在获取输入参数...")
            
            # 获取参数
            sec_type = self.sec_type_combo.currentText()
            b = float(self.b_input.text())
            h = float(self.h_input.text())
            bf = float(self.bf_input.text()) if sec_type == "T形" else 0
            hf = float(self.hf_input.text()) if sec_type == "T形" else 0
            
            # 如果T形截面的bf或hf为0，自动转为矩形截面处理
            if sec_type == "T形" and (bf <= 0 or hf <= 0):
                sec_type = "矩形"
                bf = 0
                hf = 0
            fcuk = float(self.fcuk_input.text())
            fy_grade = self.fy_combo.currentText()
            fyc_grade = self.fyc_combo.currentText()  # 使用受压钢筋强度等级控件的值
            ast = float(self.ast_input.text())
            as_t = float(self.as_t_input.text())
            asc = float(self.asc_input.text())
            as_c = float(self.as_c_input.text())
            M = float(self.m_input.text())
            is_seismic = 1 if self.seismic_combo.currentText() == "是" else 0
            gamma0 = float(self.gamma0_input.text())
            
            # 构建计算参数
            # 使用当前选中的截面索引来获取正确的序号和编号
            if self.current_section_index >= 0 and self.current_section_index < len(self.section_data):
                index = self.current_section_index
                data = self.section_data[index]
                sec_num = data.get("截面编号", data.get("sec_num", f"截面{index+1}"))
            else:
                index = 0
                sec_num = "单个计算"
            
            item = {
                "sec_num": sec_num,
                "sec_type": sec_type,
                "γ0": gamma0,
                "M": M,
                "is_seismic": is_seismic,
                "calc_params": [
                    b, h, bf, hf,
                    fcuk, fy_grade, fyc_grade,
                    ast, as_t,
                    asc, as_c, gamma0
                ]
            }
            
            # 调用计算函数，使用当前选中的截面索引作为计算序号
            self.status_bar.showMessage("正在计算...")
            x, Mu, M_out, rs_ratio, report, error_msg = calculate_single_item(item, self.current_section_index, 1)
            
            if error_msg:
                self.status_bar.showMessage(f"计算错误: {error_msg}")
                self.result_text.append(f"计算错误: {error_msg}\n")
            else:
                # 显示结果
                self.result_text.append(report + "\n")
                
                # 在左侧状态栏显示临时提示
                self.status_bar.showMessage("计算完成")
                
                # 在右侧永久状态栏区域显示计算结果
                # 确定抗弯承载力的显示名称
                mu_label = "MuE" if is_seismic == 1 else "Mu"
                mu_value = M_out if is_seismic == 1 else Mu
                
                # 构建计算结果消息，包含截面编号
                result_msg = f"编号: {sec_num} | x={x:.2f}mm | {mu_label}={mu_value:.2f}kN·m | R/S={rs_ratio:.4f}"
                self.result_status_label.setText(result_msg)
                
                # 生成结果文件（如果勾选了输出结果文件）
                if hasattr(self, 'output_result_var') and self.output_result_var.isChecked():
                    # 获取结果文件名和路径
                    result_filename = self.result_file_input.text()
                    data_file = self.file_input.text()
                    data_dir = os.path.dirname(data_file)
                    
                    # 生成计算书文件（.out格式）
                    report_result_file = os.path.join(data_dir, f"{result_filename}.out")
                    with open(report_result_file, "w", encoding="utf-8") as f:
                        f.write(self.result_text.toPlainText())
                    
                    # 显示保存成功信息
                    save_msg = f"计算书已保存到: {report_result_file}"
                    self.status_bar.showMessage(save_msg)
                    self.result_text.append(f"\n{save_msg}\n")
                
                # 滚动到顶部
                self.result_text.verticalScrollBar().setValue(0)
                
        except ValueError as e:
            error_msg = f"参数输入错误: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
        except Exception as e:
            error_msg = f"计算失败: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
    
    def calculate_batch(self):
        """批量计算"""
        try:
            # 清空文本输出框
            self.result_text.clear()
            
            data_file = self.file_input.text()
            result_file = self.result_file_input.text()
            
            if not os.path.exists(data_file):
                error_msg = f"数据文件不存在: {data_file}"
                self.status_bar.showMessage(error_msg)
                self.result_text.append(error_msg + "\n")
                return
            
            self.status_bar.showMessage("正在读取数据...")
            
            # 读取并准备数据
            df = pd.read_excel(data_file)
            from concrete.core.beam_utils import prepare_calculation_data
            param, result_data = prepare_calculation_data(df)
            
            total_count = len(param)
            self.status_bar.showMessage(f"正在计算 {total_count} 个截面...")
            
            # 保存所有报告
            all_reports = []
            error_count = 0
            
            for idx, item in enumerate(param):
                # 调用单个计算函数
                x, Mu, M_out, rs_ratio, report, error_msg = calculate_single_item(item, idx, total_count)
                
                # 添加报告到列表
                all_reports.append(report)
                
                # 记录错误
                if error_msg:
                    error_count += 1
                
                # 更新状态
                self.status_bar.showMessage(f"正在计算截面 {idx+1}/{total_count}")
                
                # 显示进度
                QApplication.processEvents()
            
            # 显示所有报告
            self.result_text.append("=====批量计算结果=====\n")
            self.result_text.append(f"数据文件: {data_file}\n")
            self.result_text.append(f"共 {total_count} 个截面\n\n")
            
            for report in all_reports:
                self.result_text.append(report + "\n")
            
            # 生成结果文件（如果勾选了输出结果文件）
            if hasattr(self, 'output_result_var') and self.output_result_var.isChecked():
                # 获取结果文件名和路径
                result_filename = self.result_file_input.text()
                data_dir = os.path.dirname(data_file)
                
                # 生成Excel文件
                excel_result_file = os.path.join(data_dir, f"{result_filename}.xlsx")
                from concrete.core.beam_utils import save_excel_result_with_style
                save_excel_result_with_style(result_data, excel_result_file, data_file)
                
                # 生成计算书文件（.out格式）
                report_result_file = os.path.join(data_dir, f"{result_filename}.out")
                with open(report_result_file, "w", encoding="utf-8") as f:
                    f.write(self.result_text.toPlainText())
                
                # 显示保存成功信息
                save_msg = f"结果已保存到: {data_dir}，Excel文件: {result_filename}.xlsx，计算书文件: {result_filename}.out"
                self.status_bar.showMessage(save_msg)
            
            # 显示总结
            summary = f"\n=====计算总结=====\n"
            summary += f"总截面数: {total_count}\n"
            summary += f"成功计算: {total_count - error_count}\n"
            summary += f"计算失败: {error_count}\n"
            if hasattr(self, 'output_result_var') and self.output_result_var.isChecked():
                data_dir = os.path.dirname(data_file)
                summary += f"结果已保存到: {data_dir}\n"
                summary += f"Excel文件: {result_filename}.xlsx\n"
                summary += f"计算书文件: {result_filename}.out\n"
            self.result_text.append(summary)
            
            # 滚动到顶部
            self.result_text.verticalScrollBar().setValue(0)
            
            # 在左侧状态栏显示临时提示
            status_msg = f"批量计算完成 | 总截面数: {total_count} | 成功: {total_count - error_count} | 失败: {error_count}"
            self.status_bar.showMessage(status_msg)
            
            # 在右侧永久状态栏区域显示批量计算总结
            result_msg = f"批量计算完成 | 构件数: {total_count} | 成功: {total_count - error_count} | 失败: {error_count}"
            self.result_status_label.setText(result_msg)
            
        except FileNotFoundError as e:
            error_msg = f"文件不存在: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
        except PermissionError as e:
            error_msg = f"权限不足: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
        except Exception as e:
            error_msg = f"批量计算失败: {str(e)}"
            self.status_bar.showMessage(error_msg)
            self.result_text.append(error_msg + "\n")
    

    

def main():
    """主函数"""
    app = QApplication(sys.argv)
    
    # 创建并显示主窗口
    window = BeamCalculationGUI()
    window.show()
    
    # 运行应用程序
    sys.exit(app.exec())

if __name__ == "__main__":
    main()